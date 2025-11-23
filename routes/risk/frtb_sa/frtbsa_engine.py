"""
FRTB-SA Capital Calculation Engine
Implements Basel Committee FRTB Standardized Approach (BCBS-D457)
Version: 1.0
Author: FRTB Risk Analytics Team
"""

import pandas as pd
import numpy as np
from datetime import datetime
import json
import warnings
from typing import Dict, List, Tuple, Optional
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')

class FRTBSAConfig:
    """Configuration class for FRTB-SA regulatory parameters"""

    # Risk weights by risk class (as per BCBS-D457)
    RISK_WEIGHTS = {
        'GIRR': {
            'major_currencies': ['USD', 'EUR', 'GBP', 'AUD', 'CAD', 'SEK', 'JPY'],
            'delta_weights': {
                '0.25Y': 0.017, '0.5Y': 0.017, '1Y': 0.016, '2Y': 0.013,
                '3Y': 0.012, '5Y': 0.011, '10Y': 0.011, '15Y': 0.011,
                '20Y': 0.011, '30Y': 0.011
            },
            'vega_weight': 1.0
        },
        'CSR_NON_SEC': {
            'buckets': {
                1: {'sector': 'IG_Sovereigns', 'weight': 0.005},
                2: {'sector': 'IG_Financials', 'weight': 0.005},
                3: {'sector': 'IG_Corporates', 'weight': 0.005},
                4: {'sector': 'HY_Sovereigns', 'weight': 0.020},
                5: {'sector': 'HY_Financials', 'weight': 0.030},
                6: {'sector': 'HY_Corporates', 'weight': 0.030},
                7: {'sector': 'Others', 'weight': 0.050}
            }
        },
        'EQUITY': {
            'buckets': {
                1: {'market_cap': 'Large_Cap', 'economy': 'Emerging', 'weight': 0.25},
                2: {'market_cap': 'Large_Cap', 'economy': 'Advanced', 'weight': 0.20},
                3: {'market_cap': 'Small_Cap', 'economy': 'Emerging', 'weight': 0.35},
                4: {'market_cap': 'Small_Cap', 'economy': 'Advanced', 'weight': 0.30},
                5: {'market_cap': 'Indexes', 'economy': 'All', 'weight': 0.15}
            }
        },
        'FX': {'weight': 0.15},
        'COMMODITY': {
            'buckets': {
                1: {'type': 'Energy_Solid', 'weight': 0.30},
                2: {'type': 'Energy_Liquid', 'weight': 0.35},
                3: {'type': 'Energy_Electric', 'weight': 0.60},
                4: {'type': 'Freight', 'weight': 0.80},
                5: {'type': 'Metals_Non_Precious', 'weight': 0.40},
                6: {'type': 'Gaseous', 'weight': 0.45},
                7: {'type': 'Precious_Metals', 'weight': 0.20},
                8: {'type': 'Grains_Oilseeds', 'weight': 0.35},
                9: {'type': 'Livestock_Dairy', 'weight': 0.25},
                10: {'type': 'Softs_Others', 'weight': 0.35}
            }
        }
    }

    # Correlation parameters
    CORRELATIONS = {
        'GIRR': {
            'tenor_correlation': 0.999,
            'curve_correlation': 0.999,
            'currency_correlation': 0.50
        },
        'CSR': {
            'name_correlation': 0.35,
            'tenor_correlation': 0.65,
            'basis_correlation': 0.999
        },
        'EQUITY': {
            'same_bucket': 0.15,
            'different_bucket': 0.0
        },
        'FX': 0.60,
        'COMMODITY': {
            'same_bucket': 0.95,
            'different_bucket': 0.20
        }
    }

class RiskFactorProcessor:
    """Process and classify risk factors according to FRTB-SA methodology"""

    def __init__(self, config: FRTBSAConfig):
        self.config = config

    def classify_risk_factor(self, row: pd.Series) -> Dict:
        """Classify a single risk factor according to FRTB-SA taxonomy"""

        risk_class = row.get('RiskClass', '')
        risk_type = row.get('Risk_Type', '')
        bucket = row.get('Bucket', '')
        qualifier = row.get('Qualifier', '')

        classification = {
            'risk_class': risk_class,
            'risk_type': risk_type,
            'bucket': bucket,
            'qualifier': qualifier,
            'label1': row.get('Label1', ''),
            'label2': row.get('Label2', ''),
            'currency': row.get('CURRENCY_1', 'USD')
        }

        # Apply specific classification rules per risk class
        if 'GIRR' in risk_class or 'General Interest' in risk_class:
            classification['risk_class'] = 'GIRR'
            classification['sub_curve'] = row.get('Curve_Type', 'OIS')

        elif 'Credit Spread' in risk_class or 'CSR' in risk_class:
            if 'secur' in risk_class.lower():
                classification['risk_class'] = 'CSR_SEC'
            else:
                classification['risk_class'] = 'CSR_NON_SEC'
            classification['credit_quality'] = row.get('Credit Quality', 'IG')
            classification['sector'] = row.get('Sector', 'Corporate')

        elif 'Equity' in risk_class or 'EQ' in risk_class:
            classification['risk_class'] = 'EQUITY'
            classification['market_cap'] = row.get('Market Cap', 'Large_Cap')
            classification['economy'] = row.get('Economy', 'Advanced')

        elif 'FX' in risk_class or 'Foreign Exchange' in risk_class:
            classification['risk_class'] = 'FX'
            classification['currency_pair'] = f"{row.get('CURRENCY_1', 'USD')}/{row.get('CURRENCY_2', 'EUR')}"

        elif 'Commodity' in risk_class or 'COMM' in risk_class:
            classification['risk_class'] = 'COMMODITY'
            classification['commodity_type'] = row.get('Comm_Type', 'Energy')

        return classification

class DeltaRiskCalculator:
    """Calculate Delta risk capital charges"""

    def __init__(self, config: FRTBSAConfig):
        self.config = config

    def calculate_weighted_sensitivity(self, sensitivity: float, risk_weight: float) -> float:
        """Calculate risk-weighted sensitivity"""
        return sensitivity * risk_weight

    def calculate_bucket_capital(self, sensitivities: pd.DataFrame, risk_class: str) -> float:
        """Calculate capital charge for a single bucket"""

        if len(sensitivities) == 0:
            return 0.0

        # Get risk weights
        if risk_class == 'GIRR':
            weights = self.config.RISK_WEIGHTS['GIRR']['delta_weights']
        elif risk_class == 'FX':
            weights = {'all': self.config.RISK_WEIGHTS['FX']['weight']}
        else:
            # Get bucket-specific weight
            bucket = sensitivities.iloc[0].get('Bucket', 1)
            if risk_class in self.config.RISK_WEIGHTS:
                bucket_config = self.config.RISK_WEIGHTS[risk_class].get('buckets', {})
                weights = {'all': bucket_config.get(bucket, {}).get('weight', 0.05)}
            else:
                weights = {'all': 0.05}

        # Calculate weighted sensitivities
        total_ws = 0
        correlations_sum = 0

        for idx, row in sensitivities.iterrows():
            # Get sensitivity amount
            sensitivity = float(row.get('FS Amount USD', 0))

            # Get appropriate weight
            if risk_class == 'GIRR':
                tenor = row.get('Label1', '1Y')
                weight = weights.get(tenor, 0.01)
            else:
                weight = weights.get('all', 0.05)

            ws = self.calculate_weighted_sensitivity(sensitivity, weight)
            total_ws += ws ** 2

            # Add correlation terms
            if risk_class in self.config.CORRELATIONS:
                corr = self.config.CORRELATIONS[risk_class]
                if isinstance(corr, dict):
                    correlations_sum += ws ** 2 * corr.get('tenor_correlation', 0.999)
                else:
                    correlations_sum += ws ** 2 * corr

        # Calculate capital charge using aggregation formula
        kb = np.sqrt(max(total_ws + correlations_sum, 0))

        return kb

class VegaRiskCalculator:
    """Calculate Vega risk capital charges"""

    def __init__(self, config: FRTBSAConfig):
        self.config = config

    def calculate_vega_capital(self, sensitivities: pd.DataFrame) -> float:
        """Calculate Vega risk capital charge"""

        vega_capital = 0.0

        for _, row in sensitivities.iterrows():
            if pd.notna(row.get('Implied_Volatility')):
                sensitivity = float(row.get('FS Amount USD', 0))
                implied_vol = float(row.get('Implied_Volatility', 0))

                # Vega capital = sensitivity * implied_vol * risk_weight
                vega_capital += abs(sensitivity * implied_vol * 1.0)

        return vega_capital

class CurvatureRiskCalculator:
    """Calculate Curvature risk capital charges"""

    def calculate_curvature_capital(self, data: pd.DataFrame) -> float:
        """Calculate Curvature risk capital charge"""

        curvature_capital = 0.0

        for _, row in data.iterrows():
            pv_up = float(row.get('PnL_Up Delta [ + RW]', 0))
            pv_down = float(row.get('PnL_Down Delta [- RW]', 0))
            pv0 = float(row.get('PV0 /MTM', 0))

            if pv0 != 0:
                # Curvature = -min(PV_up - PV0, PV_down - PV0)
                cvr_up = pv_up - pv0
                cvr_down = pv_down - pv0
                curvature = -min(cvr_up, cvr_down)
                curvature_capital += max(curvature, 0)

        return curvature_capital

class DefaultRiskCalculator:
    """Calculate Default risk capital charges"""

    def calculate_default_risk_capital(self, data: pd.DataFrame) -> Dict[str, float]:
        """Calculate Default Risk Capital charges by category"""

        drc = {
            'non_sec': 0.0,
            'sec_non_ctp': 0.0,
            'sec_ctp': 0.0
        }

        # Group by default risk categories
        for _, row in data.iterrows():
            if pd.notna(row.get('Default_Risk_Code')):
                notional = float(row.get('Trade Notional', 0))
                long_short = row.get('Long/Short', 'Long')
                multiplier = 1 if long_short == 'Long' else -1

                # Apply default risk weights (simplified)
                credit_quality = row.get('Credit Quality', 'IG')
                if credit_quality == 'IG':
                    lgd = 0.75
                else:
                    lgd = 1.0

                default_charge = abs(notional * lgd * 0.008 * multiplier)

                # Categorize
                risk_class = row.get('RiskClass', '')
                if 'CTP' in risk_class:
                    drc['sec_ctp'] += default_charge
                elif 'secur' in risk_class.lower():
                    drc['sec_non_ctp'] += default_charge
                else:
                    drc['non_sec'] += default_charge

        return drc

class ResidualRiskCalculator:
    """Calculate Residual Risk Add-On (RRAO)"""

    def calculate_rrao(self, data: pd.DataFrame) -> float:
        """Calculate Residual Risk Add-On"""

        rrao = 0.0

        for _, row in data.iterrows():
            if pd.notna(row.get('Residual Risk Class [RRC]')):
                notional = float(row.get('Trade Notional', 0))
                # RRAO = 1% of gross notional for exotic underlyings
                rrao += abs(notional) * 0.01

        return rrao

class FRTBSAEngine:
    """Main FRTB-SA calculation engine"""

    def __init__(self):
        self.config = FRTBSAConfig()
        self.rf_processor = RiskFactorProcessor(self.config)
        self.delta_calc = DeltaRiskCalculator(self.config)
        self.vega_calc = VegaRiskCalculator(self.config)
        self.curvature_calc = CurvatureRiskCalculator()
        self.default_calc = DefaultRiskCalculator()
        self.residual_calc = ResidualRiskCalculator()

        # Results storage
        self.results = {}

    def load_data(self, filepath: str) -> pd.DataFrame:
        """Load and validate input data"""

        logger.info(f"Loading data from {filepath}")

        try:
            # Support both Excel and CSV formats
            if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
                df = pd.read_excel(filepath, sheet_name=0)
            else:
                df = pd.read_csv(filepath)

            logger.info(f"Loaded {len(df)} records from {filepath}")

            # Data validation
            required_columns = ['RiskClass', 'Risk_Type', 'FS Amount USD']
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                logger.warning(f"Missing columns: {missing_cols}")

            # Convert numeric columns
            numeric_cols = ['FS Amount USD', 'Trade Notional', 'PV0 /MTM',
                          'PnL_Up Delta [ + RW]', 'PnL_Down Delta [- RW]']
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '').str.replace('(', '-').str.replace(')', ''), errors='coerce').fillna(0)

            return df

        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            raise

    def calculate_capital_charges(self, data: pd.DataFrame) -> Dict:
        """Calculate all FRTB-SA capital charges"""

        logger.info("Starting FRTB-SA capital calculation")

        capital_charges = {
            'GIRR': 0.0,
            'CSR_NON_SEC': 0.0,
            'CSR_CTP': 0.0,
            'CSR_SEC': 0.0,
            'EQUITY': 0.0,
            'COMMODITY': 0.0,
            'FX': 0.0,
            'DEFAULT_RISK_NON_SEC': 0.0,
            'DEFAULT_RISK_SEC_NON_CTP': 0.0,
            'DEFAULT_RISK_SEC_CTP': 0.0,
            'RRAO': 0.0
        }

        # Process each risk class
        risk_classes = data['RiskClass'].unique()

        for risk_class in risk_classes:
            logger.info(f"Processing risk class: {risk_class}")

            # Filter data for this risk class
            class_data = data[data['RiskClass'] == risk_class]

            # Classify risk factors
            class_data['classification'] = class_data.apply(
                lambda x: self.rf_processor.classify_risk_factor(x), axis=1
            )

            # Get standardized risk class
            std_risk_class = class_data.iloc[0]['classification']['risk_class']

            # Calculate Delta risk
            if std_risk_class in ['GIRR', 'CSR_NON_SEC', 'CSR_SEC', 'EQUITY', 'COMMODITY', 'FX']:
                # Group by bucket for aggregation
                buckets = class_data['Bucket'].unique()

                bucket_capital = 0.0
                for bucket in buckets:
                    bucket_data = class_data[class_data['Bucket'] == bucket]
                    bucket_cap = self.delta_calc.calculate_bucket_capital(bucket_data, std_risk_class)
                    bucket_capital += bucket_cap

                if std_risk_class in capital_charges:
                    capital_charges[std_risk_class] += bucket_capital

            # Calculate Vega risk (for options)
            vega_capital = self.vega_calc.calculate_vega_capital(class_data)
            if std_risk_class in capital_charges:
                capital_charges[std_risk_class] += vega_capital

        # Calculate Curvature risk
        curvature_capital = self.curvature_calc.calculate_curvature_capital(data)

        # Distribute curvature across risk classes proportionally
        total_delta_vega = sum([capital_charges[k] for k in ['GIRR', 'CSR_NON_SEC', 'EQUITY', 'COMMODITY', 'FX']])
        if total_delta_vega > 0:
            for risk_class in ['GIRR', 'CSR_NON_SEC', 'EQUITY', 'COMMODITY', 'FX']:
                capital_charges[risk_class] += curvature_capital * (capital_charges[risk_class] / total_delta_vega)

        # Calculate Default Risk
        default_charges = self.default_calc.calculate_default_risk_capital(data)
        capital_charges['DEFAULT_RISK_NON_SEC'] = default_charges['non_sec']
        capital_charges['DEFAULT_RISK_SEC_NON_CTP'] = default_charges['sec_non_ctp']
        capital_charges['DEFAULT_RISK_SEC_CTP'] = default_charges['sec_ctp']

        # Calculate Residual Risk Add-On
        capital_charges['RRAO'] = self.residual_calc.calculate_rrao(data)

        # Calculate total
        capital_charges['TOTAL'] = sum(capital_charges.values())

        logger.info(f"Total FRTB-SA Capital Charge: {capital_charges['TOTAL']:,.2f}")

        self.results = capital_charges
        return capital_charges

    def generate_qis_report(self, capital_charges: Dict, output_path: str):
        """Generate QIS format report"""

        logger.info(f"Generating QIS report: {output_path}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "QIS"

        # Define styles
        header_font = Font(bold=True, size=14)
        sub_header_font = Font(bold=True, size=12)
        cell_font = Font(size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        number_alignment = Alignment(horizontal="right")

        # Create borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Trading book QIS standardised approach'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Description
        ws.merge_cells('A3:F4')
        ws['A3'] = 'This worksheet gathers data on the standardised approach for the global trading book.'
        ws['A3'].alignment = Alignment(wrap_text=True)

        # Summary table header
        ws['A6'] = 'Summary table'
        ws['A6'].font = header_font

        ws['B8'] = 'Risk class'
        ws['B8'].font = sub_header_font
        ws['F8'] = 'Capital charge'
        ws['F8'].font = sub_header_font

        # Summary table content
        summary_data = [
            ('A) General interest rate risk (GIRR)', capital_charges.get('GIRR', 0)),
            ('B) Credit spread risk (CSR): non-securitisations', capital_charges.get('CSR_NON_SEC', 0)),
            ('C) Credit spread risk (CSR): Correlation trading portfolio', capital_charges.get('CSR_CTP', 0)),
            ('D) Credit spread risk (CSR): Securitisations (non CTP)', capital_charges.get('CSR_SEC', 0)),
            ('E) Equity risk', capital_charges.get('EQUITY', 0)),
            ('F) Commodity risk', capital_charges.get('COMMODITY', 0)),
            ('G) Foreign exchange risk', capital_charges.get('FX', 0)),
            ('H) Default risk: non-securitisations', capital_charges.get('DEFAULT_RISK_NON_SEC', 0)),
            ('I) Default risk: securitisations (non CTP)', capital_charges.get('DEFAULT_RISK_SEC_NON_CTP', 0)),
            ('J) Default risk: securitisations (CTP)', capital_charges.get('DEFAULT_RISK_SEC_CTP', 0)),
            ('K) Residual Risk Add-On', capital_charges.get('RRAO', 0))
        ]

        row = 9
        for risk_class, charge in summary_data:
            ws[f'B{row}'] = risk_class
            ws[f'F{row}'] = round(charge, 2)
            ws[f'F{row}'].number_format = '#,##0.00'
            ws[f'F{row}'].alignment = number_alignment

            # Apply borders
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].border = thin_border

            row += 1

        # Total row
        ws[f'B{row}'] = 'Total'
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'F{row}'] = round(capital_charges.get('TOTAL', 0), 2)
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].font = Font(bold=True, size=12)
        ws[f'F{row}'].alignment = number_alignment

        # Apply borders to total row
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='double')
            )

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20

        # Add detailed calculation sheets
        self._add_detailed_sheets(wb, capital_charges)

        # Save
        wb.save(output_path)
        logger.info(f"QIS report saved to {output_path}")

    def _add_detailed_sheets(self, workbook, capital_charges):
        """Add detailed calculation sheets to the workbook"""

        # Add GIRR details sheet
        ws_girr = workbook.create_sheet("GIRR_Details")
        ws_girr['A1'] = "General Interest Rate Risk - Detailed Calculations"
        ws_girr['A1'].font = Font(bold=True, size=14)

        # Add headers
        headers = ['Currency', 'Tenor', 'Sensitivity', 'Risk Weight', 'Weighted Sensitivity', 'Bucket Capital']
        for idx, header in enumerate(headers, 1):
            ws_girr.cell(row=3, column=idx, value=header).font = Font(bold=True)

        # Add placeholder for detailed data (would be populated from actual calculations)
        ws_girr['A5'] = "Detailed calculation data would be populated here"

        # Add other risk class sheets similarly...

    def generate_summary_report(self, capital_charges: Dict) -> pd.DataFrame:
        """Generate summary report as DataFrame"""

        summary_data = []

        for risk_class, charge in capital_charges.items():
            if risk_class != 'TOTAL':
                summary_data.append({
                    'Risk Class': risk_class.replace('_', ' ').title(),
                    'Capital Charge': charge,
                    'Percentage of Total': (charge / capital_charges['TOTAL'] * 100) if capital_charges['TOTAL'] > 0 else 0
                })

        df_summary = pd.DataFrame(summary_data)
        df_summary = df_summary.sort_values('Capital Charge', ascending=False)

        return df_summary

def main():
    """Main execution function"""

    print("=" * 80)
    print("FRTB-SA Capital Calculation Engine")
    print("Basel Committee on Banking Supervision - BCBS-D457")
    print("=" * 80)

    # Initialize engine
    engine = FRTBSAEngine()

    # Input and output paths
    input_file = 'FRTBSA data.xlsx'  # Change this to your input file path
    output_qis = 'FRTBSA_QIS_Output.xlsx'
    output_summary = 'FRTBSA_Summary_Report.csv'

    try:
        # Load data
        data = engine.load_data(input_file)
        print(f"\nLoaded {len(data)} records from {input_file}")

        # Calculate capital charges
        print("\nCalculating FRTB-SA capital charges...")
        capital_charges = engine.calculate_capital_charges(data)

        # Generate reports
        print("\nGenerating QIS report...")
        engine.generate_qis_report(capital_charges, output_qis)

        # Generate summary
        summary_df = engine.generate_summary_report(capital_charges)
        summary_df.to_csv(output_summary, index=False)
        print(f"Summary report saved to {output_summary}")

        # Display results
        print("\n" + "=" * 80)
        print("FRTB-SA CAPITAL CHARGES SUMMARY")
        print("=" * 80)

        for risk_class, charge in capital_charges.items():
            if risk_class != 'TOTAL':
                print(f"{risk_class.replace('_', ' ').title():.<50} {charge:>20,.2f}")

        print("-" * 80)
        print(f"{'TOTAL CAPITAL CHARGE':.<50} {capital_charges['TOTAL']:>20,.2f}")
        print("=" * 80)

        print(f"\n✓ QIS report generated: {output_qis}")
        print(f"✓ Summary report generated: {output_summary}")
        print("\nCalculation completed successfully!")

    except Exception as e:
        print(f"\n❌ Error during calculation: {str(e)}")
        logger.error(f"Calculation failed: {str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    main()
